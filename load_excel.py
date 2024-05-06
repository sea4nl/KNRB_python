from openpyxl import load_workbook
import datetime
import numpy as np  
from bokeh.plotting import figure, output_file, show 
from bokeh.models import ColumnDataSource
from tabulate import tabulate
import webbrowser
import os

class Session:
    def __init__(self, 
                 session_name, 
                 session_details,
                 session_type, 
                 minutes, 
                 kilometers = 0, 
                 harde_halen = 0):
        self.session_name = session_name
        self.session_details = session_details
        self.session_type = session_type
        self.minutes = minutes
        self.kilometers = kilometers
        self.harde_halen = harde_halen

class Day:
    def __init__(self, date):
        self.date = date
        self.sessions = []

    def add_session(self, session):
        self.sessions.append(session)

    def get_kilometers(self):
        kms = 0
        for s in self.sessions:
            kms = kms + s.kilometers
        return kms

    def get_roei_minutes(self):
        min = 0
        for s in self.sessions:
            if s.session_type == "Roeien":
                min += s.minutes
        return min

    def get_fiets_minutes(self):
        min = 0
        for s in self.sessions:
            if s.session_type == "Fietsen":
                min += s.minutes
        return min

    def name(self):
        return self.date.dayname()

class Week:
    def __init__(self, year, week_number, week_type):
        self.year = year
        self.week_number = week_number
        self.week_type = week_type
        self.days = [] 

    def add_day(self, day):
        self.days.append(day)

    def get_kilometers(self):
        kms = 0
        for d in self.days:
            kms += d.get_kilometers()
        return kms

    def get_fiets_minutes(self):
        min = 0
        for d in self.days:
            min += d.get_fiets_minutes()
        return min

    def get_roei_minutes(self):
        min = 0
        for d in self.days:
            min += d.get_roei_minutes()
        return min

class TrainingBlock:
    def __init__(self, block_name):
        self.block_name = block_name
        self.weeks = []

    def add_week(self, week):
        self.weeks.append(week)

def process_excel_file(file_path):
    try:
        # Load the Excel workbook
        workbook = load_workbook(filename=file_path, read_only=True)
        
        # Select the active sheet
        sheet = workbook.active
        
        rows = []
        # Print the values of the first few rows and columns
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=23, values_only=True):
            rows.append(row)
        # You can perform any other processing or analysis here
        return rows
        
    except FileNotFoundError:
        print("File not found. Please provide a valid file path.")
    except Exception as e:
        print("An error occurred:", e)

def process_rows(rows):
    weeks = []
    # iterate over all rows in the excel file
    for i, row in enumerate(rows):
        #check if a new week starts, it contains a row of dates
        if not isinstance(row[1], datetime.datetime):
            continue
        # create a new week
        w = Week(row[1].isocalendar()[0],row[1].isocalendar()[1],row[22])
        # iterate over all columns j
        for j, c in enumerate(row):
            # skip if it is not a day
            if not isinstance(c, datetime.datetime):
                continue
            d = Day(c)
            #find all sessions
            for ii in [0,1,2]:
                n = [0,9,18][ii]
                off = [8,8,5][ii]
                y = i+n
                if rows[y+1][j] != None:
                    print(y)
                    print(j)
                    name = rows[y+1][j]
                    print(name)
                    if name == "rust":
                        continue
                    details = rows[y+2][j]
                    print(details)
                    if "KT" in name:
                        sType = "KT"
                    elif "Fiets" in name:
                        sType = "Fietsen"
                    else:
                        sType = "Roeien"
                    if isinstance(rows[y+off][j+1], str):
                        minutes = 0
                    else:
                        minutes = rows[y+off][j+1] if rows[y+off][j+1] else 0 
                    if isinstance(rows[y+off][j], str):
                        kms = 0
                    else:
                        kms = rows[y+off][j] if rows[y+off][j] else 0 
                    if rows[y+off+1][j] == "hard":
                        hard = 0
                    else:
                        hard = rows[y+off+1][j]

                    s = Session(name,details, sType, minutes, kms, hard)
                    d.add_session(s)
            w.add_day(d)

        weeks.append(w)
    
    return weeks
# Provide the path to your Excel file
excel_file_path = "/home/sea4nl/Code/KNRB_python/Programma Stef (2021-2022).xlsx"

# Call the function to process the Excel file
rows = process_excel_file(excel_file_path)
weeks = process_rows(rows)

x = []
y_roei = []
y_fiets = []
for i, w in enumerate(weeks):
    print(w.week_number)
    x.append(i)
    y_roei.append(w.get_roei_minutes())
    y_fiets.append(w.get_fiets_minutes())

output_file("geeksforgeeks.html") 
   
p = figure(plot_width=1300, plot_height=800) 
   
# area plot 
#p.varea(x=x, y1=0, y2=y_roei,fill_color="blue") 
source = ColumnDataSource(data=dict(
    x=x,
    y1=y_roei,
    y2=y_fiets,
))

p.varea_stack(['y1', 'y2'], x='x', color=("grey", "lightgrey"), source=source) 
show(p)

head = ['Week nr','Doel','Roei kms', 'Roeiminuten', 'Fietsminuten']
weekdata = []
for w in weeks:
    weekdata.append([w.week_number, w.week_type, w.get_kilometers(), w.get_roei_minutes(), w.get_fiets_minutes()])

table=tabulate(weekdata, headers=head, tablefmt='html')

f = open('table.html','w')
f.write(table)
f.close()

filename = 'file:///'+os.getcwd()+'/' + 'table.html'
webbrowser.open_new_tab(filename)